"""Finance API routes.

Endpoints:
    GET    /                    — List invoices with filters
    POST   /                    — Create invoice (auth required)
    GET    /invoices/export      — Export invoices as Excel
    GET    /payments             — List payments
    POST   /payments             — Create payment (auth required)
    GET    /budgets              — List budgets
    POST   /budgets              — Create budget (auth required)
    PATCH  /budgets/{id}         — Update budget (auth required)
    POST   /budgets/import/file  — Import budgets from Excel/CSV (auth required)
    GET    /budgets/export       — Export budgets as Excel
    GET    /evm                  — List EVM snapshots
    POST   /evm/snapshot         — Create EVM snapshot (auth required)
    GET    /{id}                — Get single invoice
    PATCH  /{id}                — Update invoice (auth required)
    POST   /{id}/approve        — Approve invoice (auth required)
    POST   /{id}/pay            — Mark invoice as paid (auth required)

NOTE: Fixed-path routes (/payments, /budgets, /evm, /invoices/export) are
registered BEFORE the parametric /{invoice_id} route so that FastAPI does not
try to parse those path segments as UUIDs.
"""

import csv
import io
import logging
import uuid
from typing import Any

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy import select

from app.dependencies import CurrentUserId, SessionDep
from app.modules.finance.models import Invoice, ProjectBudget
from app.modules.finance.schemas import (
    BudgetCreate,
    BudgetListResponse,
    BudgetResponse,
    BudgetUpdate,
    EVMListResponse,
    EVMSnapshotCreate,
    EVMSnapshotResponse,
    InvoiceCreate,
    InvoiceListResponse,
    InvoiceResponse,
    InvoiceUpdate,
    PaymentCreate,
    PaymentListResponse,
    PaymentResponse,
)
from app.modules.finance.service import FinanceService

router = APIRouter()
logger = logging.getLogger(__name__)


def _get_service(session: SessionDep) -> FinanceService:
    return FinanceService(session)


# ── Invoices (list / create) ───────────────────────────────────────────────


@router.get(
    "/",
    response_model=InvoiceListResponse,
    summary="List invoices",
    description="Retrieve a paginated list of invoices with optional filters by project, "
    "direction (payable/receivable), and status.",
)
async def list_invoices(
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    project_id: uuid.UUID | None = Query(default=None),
    direction: str | None = Query(default=None),
    status: str | None = Query(default=None),
    offset: int = Query(default=0, ge=0),
    limit: int = Query(default=50, ge=1, le=100),
    service: FinanceService = Depends(_get_service),
) -> InvoiceListResponse:
    """List invoices with optional filters."""
    items, total = await service.list_invoices(
        project_id=project_id,
        direction=direction,
        invoice_status=status,
        offset=offset,
        limit=limit,
    )
    return InvoiceListResponse(
        items=[InvoiceResponse.model_validate(i) for i in items],
        total=total,
        offset=offset,
        limit=limit,
    )


@router.post(
    "/",
    response_model=InvoiceResponse,
    status_code=201,
    summary="Create invoice",
    description="Create a new invoice with optional line items. Set invoice_direction "
    "to 'payable' (vendor invoices) or 'receivable' (client invoices).",
)
async def create_invoice(
    data: InvoiceCreate,
    user_id: CurrentUserId,
    service: FinanceService = Depends(_get_service),
) -> InvoiceResponse:
    """Create a new invoice."""
    invoice = await service.create_invoice(data, user_id=user_id)
    return InvoiceResponse.model_validate(invoice)


# ── Export invoices as Excel ────────────────────────────────────────────────


@router.get(
    "/invoices/export",
    summary="Export invoices as Excel",
    description="Download invoices for a project as an Excel (.xlsx) file. "
    "Optionally filter by direction (payable/receivable).",
    response_description="Excel file stream (application/vnd.openxmlformats-officedocument.spreadsheetml.sheet)",
)
async def export_invoices(
    session: SessionDep,
    _user_id: CurrentUserId,
    project_id: uuid.UUID = Query(...),
    direction: str | None = Query(default=None),
) -> StreamingResponse:
    """Export invoices for a project as Excel file."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    stmt = select(Invoice).where(Invoice.project_id == project_id)
    if direction:
        stmt = stmt.where(Invoice.invoice_direction == direction)
    stmt = stmt.limit(50000)

    result = await session.execute(stmt)
    items = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"

    headers = [
        "Invoice #",
        "Direction",
        "Date",
        "Due Date",
        "Vendor/Client",
        "Subtotal",
        "Tax",
        "Total",
        "Status",
    ]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = Font(bold=True)

    for row_idx, inv in enumerate(items, 2):
        ws.cell(row=row_idx, column=1, value=inv.invoice_number)
        ws.cell(row=row_idx, column=2, value=inv.invoice_direction)
        ws.cell(row=row_idx, column=3, value=inv.invoice_date)
        ws.cell(row=row_idx, column=4, value=inv.due_date)
        ws.cell(row=row_idx, column=5, value=inv.contact_id or "")
        try:
            ws.cell(row=row_idx, column=6, value=float(inv.amount_subtotal))
        except (ValueError, TypeError):
            ws.cell(row=row_idx, column=6, value=0)
        try:
            ws.cell(row=row_idx, column=7, value=float(inv.tax_amount))
        except (ValueError, TypeError):
            ws.cell(row=row_idx, column=7, value=0)
        try:
            ws.cell(row=row_idx, column=8, value=float(inv.amount_total))
        except (ValueError, TypeError):
            ws.cell(row=row_idx, column=8, value=0)
        ws.cell(row=row_idx, column=9, value=inv.status)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="invoices_export.xlsx"'},
    )


# ── Payments (MUST be before /{invoice_id}) ─────────────────────────────────


@router.get(
    "/payments",
    response_model=PaymentListResponse,
    summary="List payments",
    description="Retrieve a paginated list of payments, optionally filtered by invoice.",
)
async def list_payments(
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    invoice_id: uuid.UUID | None = Query(default=None),
    offset: int = Query(default=0, ge=0),
    limit: int = Query(default=50, ge=1, le=100),
    service: FinanceService = Depends(_get_service),
) -> PaymentListResponse:
    """List payments with optional invoice filter."""
    items, total = await service.list_payments(
        invoice_id=invoice_id, limit=limit, offset=offset
    )
    return PaymentListResponse(
        items=[PaymentResponse.model_validate(p) for p in items],
        total=total,
    )


@router.post(
    "/payments",
    response_model=PaymentResponse,
    status_code=201,
    summary="Create payment",
    description="Record a payment against an invoice. Updates the invoice's paid amount.",
)
async def create_payment(
    data: PaymentCreate,
    user_id: CurrentUserId,
    service: FinanceService = Depends(_get_service),
) -> PaymentResponse:
    """Record a payment against an invoice."""
    payment = await service.create_payment(data)
    return PaymentResponse.model_validate(payment)


# ── Budgets (MUST be before /{invoice_id}) ──────────────────────────────────


@router.get(
    "/budgets",
    response_model=BudgetListResponse,
    summary="List budgets",
    description="Retrieve project budget lines with optional filters by project and cost category.",
)
async def list_budgets(
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    project_id: uuid.UUID | None = Query(default=None),
    category: str | None = Query(default=None),
    service: FinanceService = Depends(_get_service),
) -> BudgetListResponse:
    """List project budgets."""
    items, total = await service.list_budgets(project_id=project_id, category=category)
    return BudgetListResponse(
        items=[BudgetResponse.model_validate(b) for b in items],
        total=total,
    )


@router.post(
    "/budgets",
    response_model=BudgetResponse,
    status_code=201,
    summary="Create budget line",
    description="Create a project budget line for a specific WBS element and cost category.",
)
async def create_budget(
    data: BudgetCreate,
    user_id: CurrentUserId,
    service: FinanceService = Depends(_get_service),
) -> BudgetResponse:
    """Create a project budget line."""
    budget = await service.create_budget(data)
    return BudgetResponse.model_validate(budget)


# ── Budget import (CSV / Excel) ─────────────────────────────────────────────

_BUDGET_COLUMN_ALIASES: dict[str, list[str]] = {
    "wbs_id": [
        "wbs_id",
        "wbs code",
        "wbs",
        "code",
        "wbs_code",
    ],
    "category": [
        "category",
        "cost category",
        "kategorie",
        "type",
    ],
    "original_budget": [
        "original_budget",
        "original budget",
        "original",
        "budget",
        "amount",
    ],
    "notes": [
        "notes",
        "note",
        "remarks",
        "bemerkung",
    ],
}

_ALLOWED_BUDGET_CATEGORIES = {
    "labor",
    "material",
    "equipment",
    "subcontractor",
    "overhead",
    "contingency",
    "other",
}


def _match_budget_column(header: str) -> str | None:
    """Match a header string to a canonical column name using the alias map."""
    normalised = header.strip().lower()
    for canonical, aliases in _BUDGET_COLUMN_ALIASES.items():
        if normalised in aliases:
            return canonical
    return None


def _safe_decimal_str(value: Any, default: str = "0") -> str:
    """Parse a value to a decimal string, returning default on failure."""
    if value is None:
        return default
    if isinstance(value, (int, float)):
        return str(value)
    text = str(value).strip()
    if not text:
        return default
    # Handle European-style numbers: "1.234,56" -> "1234.56"
    if "," in text and "." in text:
        last_comma = text.rfind(",")
        last_dot = text.rfind(".")
        if last_comma > last_dot:
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        float(text)  # validate
        return text
    except (ValueError, TypeError):
        return default


def _parse_budget_rows_from_csv(content_bytes: bytes) -> list[dict[str, Any]]:
    """Parse rows from a CSV file for budget import."""
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = content_bytes.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError("Unable to decode CSV file -- unsupported encoding")

    sniffer = csv.Sniffer()
    try:
        dialect = sniffer.sniff(text[:4096], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel  # type: ignore[assignment]

    reader = csv.reader(io.StringIO(text), dialect)
    raw_headers = next(reader, None)
    if not raw_headers:
        raise ValueError("CSV file is empty or has no header row")

    column_map: dict[int, str] = {}
    for idx, hdr in enumerate(raw_headers):
        canonical = _match_budget_column(hdr)
        if canonical:
            column_map[idx] = canonical

    rows: list[dict[str, Any]] = []
    for raw_row in reader:
        row: dict[str, Any] = {}
        for idx, val in enumerate(raw_row):
            canonical = column_map.get(idx)
            if canonical:
                row[canonical] = val.strip() if isinstance(val, str) else val
        if row:
            rows.append(row)

    return rows


def _parse_budget_rows_from_excel(content_bytes: bytes) -> list[dict[str, Any]]:
    """Parse rows from an Excel (.xlsx) file for budget import."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content_bytes), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError("Excel file has no worksheets")

    rows_iter = ws.iter_rows(values_only=True)
    raw_headers = next(rows_iter, None)
    if not raw_headers:
        raise ValueError("Excel file is empty or has no header row")

    column_map: dict[int, str] = {}
    for idx, hdr in enumerate(raw_headers):
        if hdr is not None:
            canonical = _match_budget_column(str(hdr))
            if canonical:
                column_map[idx] = canonical

    rows: list[dict[str, Any]] = []
    for raw_row in rows_iter:
        row: dict[str, Any] = {}
        for idx, val in enumerate(raw_row):
            canonical = column_map.get(idx)
            if canonical and val is not None:
                row[canonical] = val
        if row:
            rows.append(row)

    wb.close()
    return rows


@router.post(
    "/budgets/import/file",
    summary="Import budgets from file",
    description="Upload an Excel (.xlsx) or CSV (.csv) file to bulk-import budget lines. "
    "Column headers are auto-detected using flexible aliases (EN/DE). "
    "Returns a summary with imported, skipped, and error counts per row.",
)
async def import_budgets_file(
    _user_id: CurrentUserId,
    project_id: uuid.UUID = Query(...),
    file: UploadFile = File(..., description="Excel (.xlsx) or CSV (.csv) file"),
    service: FinanceService = Depends(_get_service),
) -> dict[str, Any]:
    """Import project budgets from an Excel or CSV file upload.

    Expected columns:
    - **WBS Code** -- work breakdown structure code
    - **Category** -- budget category (labor, material, equipment, etc.)
    - **Original Budget** -- original budget amount
    - **Notes** -- optional notes

    Returns:
        Summary with counts of imported, skipped, and error details per row.
    """
    # Validate file type
    filename = (file.filename or "").lower()
    if not filename.endswith((".xlsx", ".csv", ".xls")):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Unsupported file type. Please upload an Excel (.xlsx) or CSV (.csv) file.",
        )

    # Read file content
    content = await file.read()
    if not content:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Uploaded file is empty.",
        )

    # Limit file size (10 MB)
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="File too large. Maximum size is 10 MB.",
        )

    # Parse rows based on file type
    try:
        if filename.endswith(".xlsx") or filename.endswith(".xls"):
            rows = _parse_budget_rows_from_excel(content)
        else:
            rows = _parse_budget_rows_from_csv(content)
    except ValueError as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Failed to parse file: {exc}",
        )
    except Exception as exc:
        logger.exception("Unexpected error parsing budget import file: %s", exc)
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Failed to parse file. Please check the format and try again.",
        )

    if not rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No data rows found in file. Check that the first row contains column headers.",
        )

    # Convert rows to BudgetCreate objects and import
    imported_count = 0
    skipped = 0
    errors: list[dict[str, Any]] = []

    for row_idx, row in enumerate(rows, start=2):
        try:
            wbs_id = str(row.get("wbs_id", "")).strip() or None

            # Parse category
            category = str(row.get("category", "")).strip().lower() or None
            if category and category not in _ALLOWED_BUDGET_CATEGORIES:
                errors.append({
                    "row": row_idx,
                    "error": (
                        f"Invalid category: '{category}'. "
                        f"Allowed: {', '.join(sorted(_ALLOWED_BUDGET_CATEGORIES))}"
                    ),
                    "data": {k: str(v)[:100] for k, v in row.items()},
                })
                continue

            # Parse amount
            original_budget = _safe_decimal_str(row.get("original_budget"))

            # Validate amount is a valid number
            try:
                float(original_budget)
            except (ValueError, TypeError):
                errors.append({
                    "row": row_idx,
                    "error": f"Invalid budget amount: {row.get('original_budget')}",
                    "data": {k: str(v)[:100] for k, v in row.items()},
                })
                continue

            # Skip rows with no data
            if not wbs_id and not category and original_budget == "0":
                skipped += 1
                continue

            data = BudgetCreate(
                project_id=project_id,
                wbs_id=wbs_id,
                category=category,
                original_budget=original_budget,
                revised_budget=original_budget,  # default revised = original
            )
            await service.create_budget(data)
            imported_count += 1

        except Exception as exc:
            errors.append({
                "row": row_idx,
                "error": str(exc),
                "data": {k: str(v)[:100] for k, v in row.items()},
            })
            logger.warning("Budget import error at row %d: %s", row_idx, exc)

    logger.info(
        "Budget file import complete: imported=%d, skipped=%d, errors=%d",
        imported_count,
        skipped,
        len(errors),
    )

    return {
        "imported": imported_count,
        "skipped": skipped,
        "errors": errors,
        "total_rows": len(rows),
    }


# ── Export budgets as Excel ──────────────────────────────────────────────────


@router.get(
    "/budgets/export",
    summary="Export budgets as Excel",
    description="Download budgets for a project as an Excel (.xlsx) file with "
    "original, revised, committed, actual, forecast, and variance columns.",
    response_description="Excel file stream (application/vnd.openxmlformats-officedocument.spreadsheetml.sheet)",
)
async def export_budgets(
    session: SessionDep,
    _user_id: CurrentUserId,
    project_id: uuid.UUID = Query(...),
) -> StreamingResponse:
    """Export budgets for a project as Excel file."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    result = await session.execute(
        select(ProjectBudget).where(ProjectBudget.project_id == project_id).limit(50000)
    )
    items = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Budgets"

    headers = [
        "WBS",
        "Category",
        "Original",
        "Revised",
        "Committed",
        "Actual",
        "Forecast",
        "Variance",
    ]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = Font(bold=True)

    for row_idx, b in enumerate(items, 2):
        ws.cell(row=row_idx, column=1, value=b.wbs_id or "")
        ws.cell(row=row_idx, column=2, value=b.category or "")
        try:
            original = float(b.original_budget)
        except (ValueError, TypeError):
            original = 0.0
        try:
            revised = float(b.revised_budget)
        except (ValueError, TypeError):
            revised = 0.0
        try:
            committed = float(b.committed)
        except (ValueError, TypeError):
            committed = 0.0
        try:
            actual = float(b.actual)
        except (ValueError, TypeError):
            actual = 0.0
        try:
            forecast = float(b.forecast_final)
        except (ValueError, TypeError):
            forecast = 0.0
        variance = revised - actual

        ws.cell(row=row_idx, column=3, value=original)
        ws.cell(row=row_idx, column=4, value=revised)
        ws.cell(row=row_idx, column=5, value=committed)
        ws.cell(row=row_idx, column=6, value=actual)
        ws.cell(row=row_idx, column=7, value=forecast)
        ws.cell(row=row_idx, column=8, value=variance)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="budgets_export.xlsx"'},
    )


@router.patch(
    "/budgets/{budget_id}",
    response_model=BudgetResponse,
    summary="Update budget line",
    description="Partially update a budget line. Only provided fields are modified.",
)
async def update_budget(
    budget_id: uuid.UUID,
    data: BudgetUpdate,
    user_id: CurrentUserId,
    service: FinanceService = Depends(_get_service),
) -> BudgetResponse:
    """Update a budget line."""
    budget = await service.update_budget(budget_id, data)
    return BudgetResponse.model_validate(budget)


# ── EVM (MUST be before /{invoice_id}) ──────────────────────────────────────


@router.get(
    "/evm",
    response_model=EVMListResponse,
    summary="List EVM snapshots",
    description="List Earned Value Management snapshots for a project. "
    "Each snapshot captures PV, EV, AC, SPI, CPI, and EAC at a point in time.",
)
async def list_evm_snapshots(
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    project_id: uuid.UUID | None = Query(default=None),
    service: FinanceService = Depends(_get_service),
) -> EVMListResponse:
    """List EVM snapshots for a project."""
    items, total = await service.list_evm_snapshots(project_id=project_id)
    return EVMListResponse(
        items=[EVMSnapshotResponse.model_validate(s) for s in items],
        total=total,
    )


@router.post(
    "/evm/snapshot",
    response_model=EVMSnapshotResponse,
    status_code=201,
    summary="Create EVM snapshot",
    description="Capture a new Earned Value Management snapshot for a project. "
    "Records planned value (PV), earned value (EV), actual cost (AC), and derived indices.",
)
async def create_evm_snapshot(
    data: EVMSnapshotCreate,
    user_id: CurrentUserId,
    service: FinanceService = Depends(_get_service),
) -> EVMSnapshotResponse:
    """Create an EVM snapshot."""
    snapshot = await service.create_evm_snapshot(data)
    return EVMSnapshotResponse.model_validate(snapshot)


# ── Finance Dashboard ─────────────────────────────────────────────────────────


@router.get(
    "/dashboard",
    summary="Get finance dashboard",
    description="Aggregated finance KPIs: payable, receivable, overdue totals, "
    "budget utilisation, cash flow overview, and budget warning level "
    "(normal / caution at 80%+ / critical at 95%+). "
    "Optionally scope to a single project via project_id query parameter.",
)
async def finance_dashboard(
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    project_id: uuid.UUID | None = Query(default=None),
    service: FinanceService = Depends(_get_service),
) -> dict:
    """Aggregated finance KPIs: payable, receivable, overdue, budget, cash flow.

    Optionally scope to a single project via ``project_id`` query parameter.
    Returns budget warning level ("normal", "caution" at 80%+, "critical" at 95%+).
    """
    return await service.get_dashboard(project_id=project_id)


# ── Invoice by ID (parametric routes LAST) ──────────────────────────────────


@router.get(
    "/{invoice_id}",
    response_model=InvoiceResponse,
    summary="Get invoice",
    description="Retrieve a single invoice by its UUID, including line items and payment history.",
)
async def get_invoice(
    invoice_id: uuid.UUID,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    service: FinanceService = Depends(_get_service),
) -> InvoiceResponse:
    """Get a single invoice by ID."""
    invoice = await service.get_invoice(invoice_id)
    return InvoiceResponse.model_validate(invoice)


@router.patch(
    "/{invoice_id}",
    response_model=InvoiceResponse,
    summary="Update invoice",
    description="Partially update an invoice. Only provided fields are modified.",
)
async def update_invoice(
    invoice_id: uuid.UUID,
    data: InvoiceUpdate,
    user_id: CurrentUserId,
    service: FinanceService = Depends(_get_service),
) -> InvoiceResponse:
    """Update an invoice."""
    invoice = await service.update_invoice(invoice_id, data)
    return InvoiceResponse.model_validate(invoice)


@router.post(
    "/{invoice_id}/approve",
    response_model=InvoiceResponse,
    summary="Approve invoice",
    description="Transition an invoice to 'approved' status. "
    "Only invoices in 'draft' or 'submitted' status can be approved.",
)
async def approve_invoice(
    invoice_id: uuid.UUID,
    user_id: CurrentUserId,
    service: FinanceService = Depends(_get_service),
) -> InvoiceResponse:
    """Approve an invoice."""
    invoice = await service.approve_invoice(invoice_id)
    return InvoiceResponse.model_validate(invoice)


@router.post(
    "/{invoice_id}/pay",
    response_model=InvoiceResponse,
    summary="Mark invoice as paid",
    description="Transition an invoice to 'paid' status. Records the payment date.",
)
async def pay_invoice(
    invoice_id: uuid.UUID,
    user_id: CurrentUserId,
    service: FinanceService = Depends(_get_service),
) -> InvoiceResponse:
    """Mark invoice as paid."""
    invoice = await service.pay_invoice(invoice_id)
    return InvoiceResponse.model_validate(invoice)
