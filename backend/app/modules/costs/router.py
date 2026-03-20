"""Cost database API routes.

Endpoints:
    GET  /autocomplete    — Fast text autocomplete for cost items (public)
    POST /                — Create a cost item (auth required)
    GET  /                — Search cost items (public, query params)
    GET  /{item_id}       — Get cost item by ID
    PATCH /{item_id}      — Update cost item (auth required)
    DELETE /{item_id}     — Delete cost item (auth required)
    POST /bulk            — Bulk import cost items (auth required)
    POST /import/file     — Import cost items from Excel/CSV file (auth required)
"""

import csv
import io
import logging
import uuid
from pathlib import Path
from typing import Any

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, status
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession

from app.dependencies import CurrentUserId, RequirePermission, SessionDep
from app.modules.costs.schemas import (
    CostAutocompleteItem,
    CostItemCreate,
    CostItemResponse,
    CostItemUpdate,
    CostSearchQuery,
)
from app.modules.costs.service import CostItemService

router = APIRouter()


def _get_service(session: SessionDep) -> CostItemService:
    return CostItemService(session)


# ── Autocomplete ──────────────────────────────────────────────────────────


@router.get("/autocomplete", response_model=list[CostAutocompleteItem])
async def autocomplete_cost_items(
    service: CostItemService = Depends(_get_service),
    q: str = Query(..., min_length=2, max_length=200, description="Search text (min 2 chars)"),
    limit: int = Query(default=8, ge=1, le=20, description="Max results to return"),
) -> list[CostAutocompleteItem]:
    """Fast text autocomplete for cost items. Public endpoint — no auth required.

    Searches cost items by description and code (case-insensitive LIKE).
    Returns compact results suitable for an autocomplete dropdown.
    """
    query = CostSearchQuery(q=q, limit=limit, offset=0)
    items, _ = await service.search_costs(query)
    return [
        CostAutocompleteItem(
            code=item.code,
            description=item.description,
            unit=item.unit,
            rate=float(item.rate),
            classification=item.classification or {},
        )
        for item in items
    ]


# ── Create ────────────────────────────────────────────────────────────────


@router.post(
    "/",
    response_model=CostItemResponse,
    status_code=201,
    dependencies=[Depends(RequirePermission("costs.create"))],
)
async def create_cost_item(
    data: CostItemCreate,
    _user_id: CurrentUserId,
    service: CostItemService = Depends(_get_service),
) -> CostItemResponse:
    """Create a new cost item."""
    item = await service.create_cost_item(data)
    return CostItemResponse.model_validate(item)


# ── Search / List ─────────────────────────────────────────────────────────


@router.get("/", response_model=list[CostItemResponse])
async def search_cost_items(
    service: CostItemService = Depends(_get_service),
    q: str | None = Query(default=None, description="Text search on code and description"),
    unit: str | None = Query(default=None, description="Filter by unit"),
    source: str | None = Query(default=None, description="Filter by source"),
    min_rate: float | None = Query(default=None, ge=0, description="Minimum rate"),
    max_rate: float | None = Query(default=None, ge=0, description="Maximum rate"),
    limit: int = Query(default=50, ge=1, le=500),
    offset: int = Query(default=0, ge=0),
) -> list[CostItemResponse]:
    """Search cost items with optional filters. Public endpoint."""
    query = CostSearchQuery(
        q=q,
        unit=unit,
        source=source,
        min_rate=min_rate,
        max_rate=max_rate,
        limit=limit,
        offset=offset,
    )
    items, _ = await service.search_costs(query)
    return [CostItemResponse.model_validate(i) for i in items]


# ── Get by ID ─────────────────────────────────────────────────────────────


@router.get("/{item_id}", response_model=CostItemResponse)
async def get_cost_item(
    item_id: uuid.UUID,
    service: CostItemService = Depends(_get_service),
) -> CostItemResponse:
    """Get a cost item by ID."""
    item = await service.get_cost_item(item_id)
    return CostItemResponse.model_validate(item)


# ── Update ────────────────────────────────────────────────────────────────


@router.patch(
    "/{item_id}",
    response_model=CostItemResponse,
    dependencies=[Depends(RequirePermission("costs.update"))],
)
async def update_cost_item(
    item_id: uuid.UUID,
    data: CostItemUpdate,
    _user_id: CurrentUserId,
    service: CostItemService = Depends(_get_service),
) -> CostItemResponse:
    """Update a cost item."""
    item = await service.update_cost_item(item_id, data)
    return CostItemResponse.model_validate(item)


# ── Delete ────────────────────────────────────────────────────────────────


@router.delete(
    "/{item_id}",
    status_code=204,
    dependencies=[Depends(RequirePermission("costs.delete"))],
)
async def delete_cost_item(
    item_id: uuid.UUID,
    _user_id: CurrentUserId,
    service: CostItemService = Depends(_get_service),
) -> None:
    """Soft-delete a cost item."""
    await service.delete_cost_item(item_id)


# ── Bulk import ───────────────────────────────────────────────────────────


@router.post(
    "/bulk",
    response_model=list[CostItemResponse],
    status_code=201,
    dependencies=[Depends(RequirePermission("costs.create"))],
)
async def bulk_import_cost_items(
    data: list[CostItemCreate],
    _user_id: CurrentUserId,
    service: CostItemService = Depends(_get_service),
) -> list[CostItemResponse]:
    """Bulk import cost items. Skips duplicates by code."""
    items = await service.bulk_import(data)
    return [CostItemResponse.model_validate(i) for i in items]


# ── File import (CSV / Excel) ────────────────────────────────────────────

logger = logging.getLogger(__name__)

# Column name aliases for flexible matching (all lowercased)
_COST_COLUMN_ALIASES: dict[str, list[str]] = {
    "code": [
        "code", "item code", "cost code", "artikelnummer", "art.nr.",
        "item", "nr", "nr.", "no", "no.", "#", "id", "position",
    ],
    "description": [
        "description", "beschreibung", "desc", "text", "bezeichnung",
        "item description", "name", "title",
    ],
    "unit": [
        "unit", "einheit", "me", "uom", "unit of measure", "measure",
    ],
    "rate": [
        "rate", "price", "cost", "unit rate", "unit price", "unit cost",
        "ep", "einheitspreis", "preis", "amount", "value",
    ],
    "currency": [
        "currency", "währung", "curr", "cur",
    ],
    "classification": [
        "classification", "din 276", "din276", "kg", "cost group",
        "nrm", "masterformat", "class", "category", "group",
    ],
}


def _match_cost_column(header: str) -> str | None:
    """Match a header string to a canonical column name using the alias map.

    Args:
        header: Raw column header text from the uploaded file.

    Returns:
        Canonical column key or None if unrecognised.
    """
    normalised = header.strip().lower()
    for canonical, aliases in _COST_COLUMN_ALIASES.items():
        if normalised in aliases:
            return canonical
    return None


def _safe_float(value: Any, default: float = 0.0) -> float:
    """Parse a value to float, returning *default* on failure.

    Handles strings with comma decimal separators (e.g. "1.234,56" -> 1234.56).
    """
    if value is None:
        return default
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return default
    # Handle European-style numbers: "1.234,56" -> "1234.56"
    if "," in text and "." in text:
        last_comma = text.rfind(",")
        last_dot = text.rfind(".")
        if last_comma > last_dot:
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return float(text)
    except (ValueError, TypeError):
        return default


def _parse_cost_rows_from_csv(content_bytes: bytes) -> list[dict[str, Any]]:
    """Parse rows from a CSV file for cost import.

    Tries UTF-8 first, then Latin-1 as fallback (common for DACH region files).
    """
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = content_bytes.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError("Unable to decode CSV file — unsupported encoding")

    sniffer = csv.Sniffer()
    try:
        dialect = sniffer.sniff(text[:4096], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel  # type: ignore[assignment]

    reader = csv.reader(io.StringIO(text), dialect)
    raw_headers = next(reader, None)
    if not raw_headers:
        raise ValueError("CSV file is empty or has no header row")

    column_map: dict[int, str] = {}
    for idx, hdr in enumerate(raw_headers):
        canonical = _match_cost_column(hdr)
        if canonical:
            column_map[idx] = canonical

    rows: list[dict[str, Any]] = []
    for raw_row in reader:
        row: dict[str, Any] = {}
        for idx, val in enumerate(raw_row):
            canonical = column_map.get(idx)
            if canonical:
                row[canonical] = val.strip() if isinstance(val, str) else val
        if row:
            rows.append(row)

    return rows


def _parse_cost_rows_from_excel(content_bytes: bytes) -> list[dict[str, Any]]:
    """Parse rows from an Excel (.xlsx) file for cost import."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content_bytes), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError("Excel file has no worksheets")

    rows_iter = ws.iter_rows(values_only=True)
    raw_headers = next(rows_iter, None)
    if not raw_headers:
        raise ValueError("Excel file is empty or has no header row")

    column_map: dict[int, str] = {}
    for idx, hdr in enumerate(raw_headers):
        if hdr is not None:
            canonical = _match_cost_column(str(hdr))
            if canonical:
                column_map[idx] = canonical

    rows: list[dict[str, Any]] = []
    for raw_row in rows_iter:
        row: dict[str, Any] = {}
        for idx, val in enumerate(raw_row):
            canonical = column_map.get(idx)
            if canonical and val is not None:
                row[canonical] = val
        if row:
            rows.append(row)

    wb.close()
    return rows


@router.post(
    "/import/file",
    dependencies=[Depends(RequirePermission("costs.create"))],
)
async def import_cost_file(
    _user_id: CurrentUserId,
    file: UploadFile = File(..., description="Excel (.xlsx) or CSV (.csv) file"),
    service: CostItemService = Depends(_get_service),
) -> dict[str, Any]:
    """Import cost items from an Excel or CSV file upload.

    Accepts a multipart file upload. The file must be .xlsx or .csv.

    Expected columns (flexible auto-detection):
    - **Code / Item Code / Nr.** — unique cost item code (required)
    - **Description / Beschreibung / Text** — description (required)
    - **Unit / Einheit / ME** — unit of measurement
    - **Rate / Price / Cost / EP** — unit rate or price
    - **Currency / Währung** — currency code (defaults to EUR)
    - **Classification / DIN 276 / KG** — classification code

    Returns:
        Summary with counts of imported, skipped, and error details per row.
    """
    # Validate file type
    filename = (file.filename or "").lower()
    if not filename.endswith((".xlsx", ".csv", ".xls")):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Unsupported file type. Please upload an Excel (.xlsx) or CSV (.csv) file.",
        )

    # Read file content
    content = await file.read()
    if not content:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Uploaded file is empty.",
        )

    # Limit file size (10 MB)
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="File too large. Maximum size is 10 MB.",
        )

    # Parse rows based on file type
    try:
        if filename.endswith(".xlsx") or filename.endswith(".xls"):
            rows = _parse_cost_rows_from_excel(content)
        else:
            rows = _parse_cost_rows_from_csv(content)
    except ValueError as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Failed to parse file: {exc}",
        )
    except Exception as exc:
        logger.exception("Unexpected error parsing cost import file: %s", exc)
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Failed to parse file. Please check the format and try again.",
        )

    if not rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No data rows found in file. Check that the first row contains column headers.",
        )

    # Convert rows to CostItemCreate objects and import via service
    items_to_import: list[CostItemCreate] = []
    skipped = 0
    errors: list[dict[str, Any]] = []
    auto_code = 1

    for row_idx, row in enumerate(rows, start=2):
        try:
            code = str(row.get("code", "")).strip()
            description = str(row.get("description", "")).strip()

            # Skip rows without both code and description
            if not code and not description:
                skipped += 1
                continue

            # Auto-generate code if missing
            if not code:
                code = f"IMPORT-{auto_code:06d}"
            auto_code += 1

            # Skip obvious summary rows
            desc_lower = description.lower()
            if desc_lower in (
                "total", "grand total", "summe", "gesamt", "gesamtsumme",
                "subtotal", "zwischensumme",
            ):
                skipped += 1
                continue

            # Parse unit (default: pcs)
            unit = str(row.get("unit", "pcs")).strip()
            if not unit:
                unit = "pcs"

            # Parse rate
            rate = _safe_float(row.get("rate"), default=0.0)

            # Parse currency (default: EUR)
            currency = str(row.get("currency", "EUR")).strip().upper()
            if not currency:
                currency = "EUR"

            # Build classification
            classification: dict[str, str] = {}
            class_value = str(row.get("classification", "")).strip()
            if class_value:
                classification["code"] = class_value

            items_to_import.append(
                CostItemCreate(
                    code=code,
                    description=description,
                    unit=unit,
                    rate=rate,
                    currency=currency,
                    source="file_import",
                    classification=classification,
                )
            )

        except Exception as exc:
            errors.append({
                "row": row_idx,
                "error": str(exc),
                "data": {k: str(v)[:100] for k, v in row.items()},
            })
            logger.warning("Cost import error at row %d: %s", row_idx, exc)

    # Bulk import via service (handles duplicate detection)
    imported_items = await service.bulk_import(items_to_import) if items_to_import else []
    imported_count = len(imported_items)
    skipped_by_duplicate = len(items_to_import) - imported_count

    logger.info(
        "Cost file import complete: imported=%d, skipped=%d (empty) + %d (duplicate), errors=%d",
        imported_count,
        skipped,
        skipped_by_duplicate,
        len(errors),
    )

    return {
        "imported": imported_count,
        "skipped": skipped + skipped_by_duplicate,
        "errors": errors,
        "total_rows": len(rows),
    }


# ── Load CWICR database from local DDC Toolkit ──────────────────────────────


CWICR_SEARCH_PATHS = [
    "../../DDC_Toolkit/pricing/data/excel",
    "../DDC_Toolkit/pricing/data/excel",
    str(Path.home() / "DDC_Toolkit" / "pricing" / "data" / "excel"),
    str(Path.home() / "Desktop" / "CodeProjects" / "DDC_Toolkit" / "pricing" / "data" / "excel"),
]


def _find_cwicr_file(db_id: str) -> Path | None:
    """Find a CWICR database file by database ID (e.g., DE_BERLIN).

    Priority: Parquet (fastest, most reliable) > Excel SIMPLE > Excel any.
    """
    # Priority 1: Parquet files (fastest and most reliable)
    for search_path in CWICR_SEARCH_PATHS:
        parquet_path = Path(search_path).parent / "parquet"
        if parquet_path.exists():
            for f in parquet_path.iterdir():
                if f.name.startswith(db_id) and f.suffix == ".parquet":
                    return f

    # Priority 2: Excel SIMPLE
    for search_path in CWICR_SEARCH_PATHS:
        p = Path(search_path)
        if not p.exists():
            continue
        for f in p.iterdir():
            if f.name.startswith(db_id) and "_SIMPLE" in f.name and f.suffix == ".xlsx":
                return f

    # Priority 3: Any Excel
    for search_path in CWICR_SEARCH_PATHS:
        p = Path(search_path)
        if not p.exists():
            continue
        for f in p.iterdir():
            if f.name.startswith(db_id) and f.suffix == ".xlsx":
                return f

    return None


@router.post(
    "/load-cwicr/{db_id}",
    dependencies=[Depends(RequirePermission("costs.create"))],
)
async def load_cwicr_database(
    db_id: str,
    session: SessionDep,
    _user_id: CurrentUserId,
) -> dict:
    """Load a CWICR regional database from local DDC Toolkit files.

    Optimized: reads Parquet, deduplicates by rate_code (55K unique items
    from 900K total rows), then bulk-inserts into SQLite.
    Typical time: 10-30 seconds.
    """
    import logging
    import time

    import pandas as pd
    from sqlalchemy import text

    logger = logging.getLogger(__name__)
    start = time.monotonic()

    # Find the file
    cwicr_path = _find_cwicr_file(db_id)
    if not cwicr_path:
        raise HTTPException(
            status_code=404,
            detail=f"CWICR database '{db_id}' not found. "
            f"Install DDC_Toolkit at ~/Desktop/CodeProjects/DDC_Toolkit",
        )

    logger.info("Loading CWICR from %s", cwicr_path)

    # Read file
    if cwicr_path.suffix == ".parquet":
        df = pd.read_parquet(cwicr_path)
    else:
        df = pd.read_excel(cwicr_path, engine="openpyxl")

    total_rows = len(df)
    logger.info("Raw data: %d rows", total_rows)

    # Normalize column names
    df.columns = [str(c).strip().lower() for c in df.columns]

    # CWICR-specific: deduplicate by rate_code to get unique work items
    if "rate_code" in df.columns:
        df = df.drop_duplicates(subset="rate_code", keep="first")
        logger.info("After dedup by rate_code: %d unique items", len(df))

    # Map CWICR columns to our schema
    col_map = {
        "code": next((c for c in df.columns if c in ("rate_code", "code", "id")), None),
        "description": next((c for c in df.columns if c in (
            "rate_original_name", "rate_final_name", "description", "name",
            "subsection_name", "section_name",
        )), None),
        "unit": next((c for c in df.columns if c in ("rate_unit", "unit")), None),
        "rate": next((c for c in df.columns if c in (
            "total_cost_per_position", "total_resource_cost_per_position",
            "cost", "price", "rate",
        )), None),
        "category": next((c for c in df.columns if c in ("collection_name", "department_name", "category")), None),
    }

    desc_col = col_map["description"]
    if not desc_col:
        raise HTTPException(400, f"No description column. Columns: {list(df.columns)[:15]}")

    # Build description: combine original_name + final_name for richer text
    if "rate_original_name" in df.columns and "rate_final_name" in df.columns:
        df["_full_desc"] = (
            df["rate_original_name"].fillna("").astype(str) + " " +
            df["rate_final_name"].fillna("").astype(str)
        ).str.strip()
        desc_col = "_full_desc"

    # Filter: must have description
    df = df[df[desc_col].astype(str).str.len() >= 3]
    logger.info("After filtering: %d items with valid descriptions", len(df))

    # Bulk insert using raw SQL for speed (10-30 seconds vs 41 minutes)
    imported = 0
    skipped = 0
    batch_size = 500
    items_to_insert = []

    for _, row in df.iterrows():
        desc = str(row.get(desc_col, ""))[:500].strip()
        code = str(row.get(col_map["code"], ""))[:100].strip() if col_map["code"] else ""
        if not code:
            code = f"CWICR-{db_id}-{imported:06d}"
        unit = str(row.get(col_map["unit"], "m2"))[:20].strip() or "m2"

        rate = 0.0
        if col_map["rate"]:
            try:
                rv = row.get(col_map["rate"], 0)
                rate = float(rv) if rv is not None and pd.notna(rv) else 0.0
            except (ValueError, TypeError):
                rate = 0.0

        category = ""
        if col_map["category"]:
            category = str(row.get(col_map["category"], ""))[:100]

        items_to_insert.append({
            "code": code,
            "description": desc,
            "unit": unit,
            "rate": str(rate),
            "currency": "",
            "source": "cwicr",
            "classification": {},
            "tags": [],
            "components": [],
            "descriptions": {},
            "is_active": True,
            "region": db_id,
        })

        if len(items_to_insert) >= batch_size:
            count = await _bulk_insert_costs(session, items_to_insert)
            imported += count
            skipped += len(items_to_insert) - count
            items_to_insert.clear()

    # Final batch
    if items_to_insert:
        count = await _bulk_insert_costs(session, items_to_insert)
        imported += count
        skipped += len(items_to_insert) - count

    duration = round(time.monotonic() - start, 1)
    logger.info("CWICR %s: %d imported, %d skipped in %.1fs", db_id, imported, skipped, duration)

    return {
        "imported": imported,
        "skipped": skipped,
        "total_rows": total_rows,
        "unique_items": len(df),
        "database": db_id,
        "source_file": cwicr_path.name,
        "duration_seconds": duration,
    }


async def _bulk_insert_costs(session: AsyncSession, items: list[dict]) -> int:
    """Bulk insert cost items, skipping duplicates by code."""
    from app.modules.costs.models import CostItem
    import uuid
    from datetime import datetime, UTC

    inserted = 0
    for item in items:
        try:
            obj = CostItem(
                id=uuid.uuid4(),
                code=item["code"],
                description=item["description"],
                unit=item["unit"],
                rate=item["rate"],
                currency=item["currency"],
                source=item["source"],
                classification=item.get("classification", {}),
                tags=item.get("tags", []),
                components=item.get("components", []),
                descriptions=item.get("descriptions", {}),
                is_active=True,
                region=item.get("region", ""),
            )
            session.add(obj)
            inserted += 1
        except Exception:
            pass

    try:
        await session.flush()
    except Exception:
        await session.rollback()
        # Try one-by-one on failure (duplicate codes)
        inserted = 0
        for item in items:
            try:
                obj = CostItem(
                    id=uuid.uuid4(),
                    code=item["code"],
                    description=item["description"],
                    unit=item["unit"],
                    rate=item["rate"],
                    currency=item["currency"],
                    source=item["source"],
                    is_active=True,
                )
                session.add(obj)
                await session.flush()
                inserted += 1
            except Exception:
                await session.rollback()

    await session.commit()
    return inserted


# ── Delete CWICR database ───────────────────────────────────────────────────


@router.delete(
    "/actions/clear-database",
    dependencies=[Depends(RequirePermission("costs.delete"))],
)
async def clear_cost_database(
    session: SessionDep,
    _user_id: CurrentUserId,
    source: str = Query(default="", description="Filter by source (e.g. 'cwicr'). Empty = delete ALL."),
) -> dict:
    """Delete cost items. Optionally filter by source."""
    from sqlalchemy import delete as sql_delete
    from app.modules.costs.models import CostItem

    if source:
        stmt = sql_delete(CostItem).where(CostItem.source == source)
    else:
        stmt = sql_delete(CostItem)

    result = await session.execute(stmt)
    await session.commit()
    count = result.rowcount  # type: ignore[union-attr]

    return {"deleted": count, "source_filter": source or "all"}


# ── Export cost database as Excel ────────────────────────────────────────────


@router.get(
    "/actions/export-excel",
    dependencies=[Depends(RequirePermission("costs.list"))],
)
async def export_cost_database(
    session: SessionDep,
    _user_id: CurrentUserId,
) -> StreamingResponse:
    """Export all cost items as Excel file."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    from app.modules.costs.models import CostItem
    from sqlalchemy import select

    result = await session.execute(select(CostItem).where(CostItem.is_active.is_(True)).limit(50000))
    items = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Cost Database"

    headers = ["Code", "Description", "Unit", "Rate", "Currency", "Source", "Region"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = Font(bold=True)

    for row_idx, item in enumerate(items, 2):
        ws.cell(row=row_idx, column=1, value=item.code)
        ws.cell(row=row_idx, column=2, value=item.description)
        ws.cell(row=row_idx, column=3, value=item.unit)
        try:
            ws.cell(row=row_idx, column=4, value=float(item.rate))
        except (ValueError, TypeError):
            ws.cell(row=row_idx, column=4, value=0)
        ws.cell(row=row_idx, column=5, value=item.currency)
        ws.cell(row=row_idx, column=6, value=item.source)
        ws.cell(row=row_idx, column=7, value=getattr(item, "region", ""))

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="cost_database.xlsx"'},
    )
